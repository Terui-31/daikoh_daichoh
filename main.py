import cv2, pyocr, glob, os, logging, openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from PIL.ExifTags import TAGS
from pprint import pprint
import warnings
os.environ['TF_CPP_MIN_LOG_LEVEL']='3'
warnings.simplefilter(action='ignore', category=FutureWarning)
import tensorflow as tf
from classifier import resnet
from detector import YOLO
from ocr_preprocess import adjust, rotate_img, binary, otsu_thresh, rgb2gray, toAlpha2
from PIL import JpegImagePlugin

JpegImagePlugin._getmp = lambda x: None
warnings.simplefilter(action='ignore', category=Warning)
tf.get_logger().setLevel('INFO')
tf.autograph.set_verbosity(0)
tf.get_logger().setLevel(logging.ERROR)


def _main():

    """
    【分類器】
    """
    classes = ['土工事', '鉄筋工事', 'コンクリート工事', '鉄骨工事', '屋根工事',  '塗装工事', '舗装工事'] #分類クラス
    book_dict = {} #辞書の初期化
    for i in range(len(classes)):
        book_dict[classes[i]] = [[], [], []] #画像パス、撮影日、テキストを保持できる形にする
    image_pathes = glob.glob('./data/*.jpg')
    for i, image_path in enumerate(image_pathes):
        print("")
        print("#"*30)
        print("###"+str(i+1)+"枚目の画像を読み込み中...")
        print("#"*30)
        print("")
        print("【画像パスを取得：{}】".format(image_path))
        label = resnet(image_path)
        print("【{}に分類】".format(label))

        """
        【EXIF】
        撮影日情報の取得
        """
        im = PILImage.open(image_path)
        exif_dict = im.getexif()
        exif = {}
        for key, value in exif_dict.items():
            exif[TAGS.get(key, key)] = value  #辞書の値を新たな辞書のキーとしている
        if "DateTime" in exif:
            datetime = exif["DateTime"]
        elif "DateTimeDigitized" in exif:
            datetime = exif["DateTimeDigitized"]
        elif "DateTimeOriginal" in exif:
            datetime = exif["DateTimeOriginal"]
        else:
            datetime = None
        if datetime is not None:
            ymd = datetime.split(" ")[0].split(":")
            ymd = "撮影日：{}年{}月{}日".format(ymd[0], ymd[1], ymd[2])
            print("【{}】".format(ymd))
        else:
            ymd = None
            print("【！撮影日情報は取得されませんでした】")

        """
        【検出器】
        PIL形式で画像そのものをdetect_imageの引数に渡し、
        黒板を検出したバウンディングボックスの位置情報をタプルで返す
        bbox = (left, top, right, bottom)
        """
        im = PILImage.open(image_path)
        try:
            bbox = YOLO.detect_image(YOLO(), im)
            print("黒板を検出:"+str(bbox))
        except:
            bbox = None
            print("【！黒板は検出されませんでした】")

        #bboxが検出された場合のみ、OCR用に画像内の黒板を切り取る
        if bbox is not None:    
            im_crop = im.crop(bbox) #黒板を切り抜く
            #im_crop.save("./tmp1_crop.jpg")

            """
            【OCR】
            """
            #前処理
            width, height = im_crop.size
            expand_rate = 1.6
            #1. 拡大 (pillow)
            im_expand = im_crop.resize((int(width*expand_rate), int(height*expand_rate)))
            im_expand.save("./out/tmp2_expand.jpg")
            #2. 角度調整 (cv2)
            im_expand = cv2.imread("./out/tmp2_expand.jpg")
            os.remove("./out/tmp2_expand.jpg")
            try:
                im_rotate = rotate_img(im_expand)
            except:
                im_rotate = im_expand
            #3. グレースケール、大津の二値化 (cv2)
            im_gray = rgb2gray(im_rotate)
            th, im_binary = otsu_thresh(im_gray)
            #4. 色彩反転（文字の黒色化、白背景化） (cv2)
            im_invert = cv2.bitwise_not(im_binary)
            cv2.imwrite("./out/tmp6_invert.jpg", im_invert)
            #pyocr のツールをロード
            tools = pyocr.get_available_tools()
            tool = tools[0]
            #テキスト抽出
            im_ = PILImage.open('./out/tmp6_invert.jpg')
            os.remove('./out/tmp6_invert.jpg')
            txt = tool.image_to_string(
                im_,
                lang = "jpn",
                builder=pyocr.builders.TextBuilder(tesseract_layout=3)
            )
            #OCRテキストの行間調整
            txt = txt.rsplit('\n\n')
            data_list_no_space = [a for a in txt if a != ' ']
            txt = '\n'.join(data_list_no_space)
            txt = txt.replace(' | ', '\n')
            print("")
            print("+"*35)
            print(txt)
            print("+"*35)
            print("")
        #bboxが検出されない場合
        else:
            txt = None
        #出力のラベルに応じて情報を辞書に保持
        for cl in classes:
            if label == cl:
                book_dict[cl][0].append(image_path)
                book_dict[cl][1].append(ymd)
                book_dict[cl][2].append(txt)

    """
    【Openpyxl】
    ラベル名ごとのkeyを持つ辞書を使いエクセルのワークブックを作成し、保持された順番で画像・テキスト等を貼り付けていく
    """
    #下線を設定
    border = Border(bottom=Side(style='thin', color='000000'))
    #各工種のブックを作成
    for l in range(len(classes)):
        if book_dict[classes[l]][0] == []:
            pass
        else:
            book = Workbook() #ブック生成
            sheet0 = book['Sheet'] #シートを取得
            sheet0.title = "Sheet0" #シート名の変更
            len_class0 = len(book_dict[classes[l]][0])
            print("{}の写真枚数：{}".format(classes[l], len_class0))
            book_counts = 0 #counts が len_class0 を超えたら処理を終えるようにする
            n_sheet = 0
            while book_counts < len_class0:
                sheet = book["Sheet{}".format(n_sheet)] #n_sheet枚目のシート
                #プリントレイアウトの微調整(デフォルトの列幅は8.38)
                for col in range(5):
                    sheet.column_dimensions['{}'.format(toAlpha2(1+(col*9)))].width = 3.0
                    sheet.column_dimensions['{}'.format(toAlpha2(2+(col*9)))].width = 13.76
                #sheet0のキャプション記載用の下線生成
                for k in range(2):
                    for j in range(5):
                        for i in range(3):
                            for row_num in range((1+k*49)+(i*16), (15+k*49)+(i*16), 1): 
                                for col_num in range(7+(j*9), 10+(j*9)):
                                    sheet.cell(row=row_num ,column=col_num).border = border
                onesheet_counts = 0 #counts が 30を超えたら次のシートを生成・移行
                #sheet0の画像を挿入
                for k in range(2):
                    for j in range(5):
                        for i in range(3):
                            img = XLImage(book_dict[classes[l]][0][i+(j*3)+(k*15)+(n_sheet*30)])
                            img.height = 247
                            img.width = 355
                            img.anchor = '{}{}'.format(toAlpha2(2+(j*9)), (2+(k*49))+(i*16))
                            sheet.add_image(img)
                            sheet["{}{}".format(toAlpha2(7+(j*9)), (1+(k*49))+(i*16))] = book_dict[classes[l]][1][i+(j*3)+(k*15)+(n_sheet*30)] #1行目に日付
                            sheet["{}{}".format(toAlpha2(7+(j*9)), (2+(k*49))+(i*16))] = classes[l] #2行目に工事種目
                            try:
                                words = book_dict[classes[l]][2][i+(j*3)+(k*15)+(n_sheet*30)].split("\n") #4行目以降に施工状況
                                for word_id, word in enumerate(words):
                                    sheet["{}{}".format(toAlpha2(7+(j*9)), ((word_id+4)+k*49)+(i*16))] = word
                            except:
                                words = None
                            book_counts += 1
                            onesheet_counts += 1
                            if book_counts == len_class0:
                                break
                            elif onesheet_counts == 30:
                                book.create_sheet() #sheet{n_sheet+1}を作成
                                n_sheet += 1
                                sheet_add = book["Sheet"]
                                sheet_add.title = "Sheet{}".format(n_sheet)
                            else:
                                continue
                            break
                        else:
                            continue
                        break
                    else:
                        continue
                    break
            #Sheet全体の罫線を消す
            sheet_names = book.get_sheet_names()
            for sheet_name in sheet_names:
                sheet = book['{}'.format(sheet_name)]
                sheet.sheet_view.showGridLines = False
            #鉄骨工事のブックを保存
            book.save('./out/{}.xlsx'.format(classes[l]))

if __name__ == '__main__':
    _main()